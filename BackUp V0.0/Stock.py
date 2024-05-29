import requests
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

def convert_text_to_values(file_path):
    # 打开 Excel 文件
    workbook = load_workbook(file_path)
    # 选择要操作的表
    sheet = workbook.active
    # 遍历单元格，将文本转换为值
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                # 尝试将文本转换为对应的数据类型
                try:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    # 检查是否为百分比文本
                    if isinstance(cell.value, str) and cell.value.endswith('%'):
                        # 如果是百分比文本，将百分号去掉，并转换为小数
                        cell.value = float(cell.value[:-1]) / 100
                        cell.number_format = '0.00%'
                    else:
                        # 否则，尝试直接转换为对应的数据类型
                        cell.value = eval(cell.value)

                    if cell.value > 0:
                        cell.font = Font(color="FF0000")  # 红色
                    elif cell.value < 0:
                        cell.font = Font(color="00FF00")  # 绿色
                except:
                    # 如果转换失败，保持文本不变
                    pass

    # 遍历所有列
    for col in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col)
        column_cells = sheet[column_letter]
        # 遍历当前列的所有单元格，找到最宽的内容
        max_length = 0
        for cell in column_cells:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
        sheet.column_dimensions[column_letter].width = max(10, max_length + 2)

    # 保存修改后的 Excel 文件
    workbook.save(file_path)
#----------------------------------------
def String2Number(s):
    try:
        if "%" in s:  # 如果字符串是百分比
            percent_str = s.strip("%")  # 去除百分比符号
            percent_float = float(percent_str) / 100  # 将字符串转换为浮点数，并除以 100
            return percent_float
        elif "-" in s:  # 如果无数据，银行股
             return float(-100)
        elif "亏损" in s:
                return float(-100) #亏损
        return float(s)
    except ValueError:
        # 如果无法将字符串转换为浮点数，则返回默认值或者抛出异常，具体处理方式取决于您的需求
        return float(-100) # 或者 raise ValueError("Invalid input: {}".format(s))

#----------------------------------------
#将数据按顺序写入到指定的Excel工作簿的一行中
def write_to_excel(workbook, data, row):
    # 选择活动的工作表
    ws = workbook.active
    # 写入数据到一行中
    for idx, value in enumerate(data, start=1):
        # 写入数据到指定的单元格
        ws.cell(row=row, column=idx, value=value)
        cell = ws.cell(row,idx)
        cell.font = Font(bold=True)

def replace_substring(original_str, target_str, replacement_str):
    return original_str.replace(target_str, replacement_str)

def adjust_column_width(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            if cell.value:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = max_length + 2  # 可以根据需要调整增加的宽度
        worksheet.column_dimensions[column].width = adjusted_width
#-------------------------------------------------------------------------

MedicalStockList = [
                "SZ300009", #安科生物 - 生物制剂 ！！!
                "SZ000661", #长春高新 !!!
                "SH688278", #特宝生物
                "SZ000739",  #普洛药业
                "SH600750",  #江中药业
                "SH600285",  #羚锐制药
                "SZ000028",  #国药一致
                "SZ002422",  #科伦药业
                "SZ300529",  #健帆生物!! 血透
                "SH688617",  #惠泰医疗!! 介入
                "SH600566",  # 济川药业
                "SZ000423",  # 东阿阿胶
                "SZ002223",  # 鱼跃医疗
                ]
MYStockList = [
             "SH600016",  #民生银行
             "SH600315",  #上海家化
             "SZ002507",  #涪陵榨菜
             "SH603868",  #飞科电器
             "SZ002223",  #鱼跃医疗
             "SZ000963",  #华东医药
             "SH600566",  #济川药业
             "SH688981",  #中芯国际
             "SZ000423",  #东阿阿胶
             "SH600309",  #万华化学
             "SZ000858",  #五粮液
            #"SZ300892",  #品渥股份
            #"SZ003030",  #祖名股份
            #"SZ000553",  #安道麦
            "SZ002511",  #中顺洁柔
            ]

HROEStockList = [
             "SZ300009", #安科生物 - 生物制剂 ！！!
             "SH603688", #石英股份- 石英材料！！！
            #"SZ300856", #科思股份 - 化妆品防晒剂！
            #"SZ301004", #嘉益股份 - 保温杯！
            #"SH603855", #华荣股份 - 防爆设备！
            #"SH600873", #梅花生物 - 味精氨基酸
            #"SH603605",#珀莱雅
            #"SH603298", #杭叉集团 - 叉车
            #"SZ002677", #浙江美大 - 集成灶
            #"SZ000895" - 双汇发展 -  肉类
            #"SZ002236" - 大华股份， 安防
            #"SZ300750" - 宁德时代， 电池
            #"SH600096", #云天化 - 磷肥头部
            #"SH603529"  # 爱玛科技 - 电动自行车
            #"SH600803", #新奥股份 - 燃气头部
            #"SH688516", #奥特维- 半导体装备
            #"SZ300274", #阳光电源 - 新能源
            #"SH601012",  隆基绿能 - N型电池
            #"SH688556", "SZ300861", #美畅股份, 高测股份 - 金刚线
          ]


with open("QualifiedStockNumber.txt", 'r') as file:
    StockList = file.read().splitlines()
    print(StockList)
#StockList = HROEStockList

TableURL = "https://xueqiu.com/S/stockID"
urlDetail = "https://xueqiu.com/snowman/S/stockID/detail#/ZYCWZB"
urlHis = "https://eniu.com/gu/stockID/market_value"

TableHead = ["  ","股价","总市值","市盈率(TTM)","市净率","股息率(TTM)","换手","量比","总股本"]
TableHeadRep = ["  ","股价","涨跌","总市值","市盈率","市净率","股息率%","换手率%","量比%","总股本","  ","  ","ROE/PE", "利润率消耗"]
TableValue = [" ",
              "净资产收益率","同比",
              "营业收入","同比",
              "毛利率","同比",
              "净利率", "同比",
              "每股现金流","同比",
              "扣非净利润","同比",
              "资产负债率",
              ]
XpathTable = ["/html/body/div[1]/div[2]/div[2]/div/div[4]/div/table/tbody/tr[15]/td[2]/p[1]",  # 净资产收益率数值
              "/html/body/div[1]/div[2]/div[2]/div/div[4]/div/table/tbody/tr[15]/td[2]/p[2]",  # 净资产收益率同比
              "/html/body/div[1]/div[2]/div[2]/div/div[4]/div/table/tbody/tr[2]/td[2]/p[1]",   # 营业收入数值
              "/html/body/div[1]/div[2]/div[2]/div/div[4]/div/table/tbody/tr[2]/td[2]/p[2]",   # 营业收入同比
              "/html/body/div[1]/div[2]/div[2]/div/div[4]/div/table/tbody/tr[19]/td[2]/p[1]",  # 毛利率数值
              "/html/body/div[1]/div[2]/div[2]/div/div[4]/div/table/tbody/tr[19]/td[2]/p[2]",  # 毛利率同比
              "/html/body/div[1]/div[2]/div[2]/div/div[4]/div/table/tbody/tr[20]/td[2]/p[1]",  # 净利率数值
              "/html/body/div[1]/div[2]/div[2]/div/div[4]/div/table/tbody/tr[20]/td[2]/p[2]",  # 净利率同比
              "/html/body/div[1]/div[2]/div[2]/div/div[4]/div/table/tbody/tr[13]/td[2]/p[1]",  # 每股现金流数值
              "/html/body/div[1]/div[2]/div[2]/div/div[4]/div/table/tbody/tr[13]/td[2]/p[2]",  # 每股现金流同比
              "/html/body/div[1]/div[2]/div[2]/div/div[4]/div/table/tbody/tr[6]/td[2]/p[1]",  # 扣非净利润数值
              "/html/body/div[1]/div[2]/div[2]/div/div[4]/div/table/tbody/tr[6]/td[2]/p[2]",  # 扣非净利润同比
              "/html/body/div[1]/div[2]/div[2]/div/div[4]/div/table/tbody/tr[22]/td[2]/p[1]",  # 资产负债率数值
              ]

TongbeCheck_xpath = "/html/body/div[1]/div[2]/div[2]/div/div[2]/div[2]/label"
year_button_xpath = "/html/body/div[1]/div[2]/div[2]/div/div[2]/div[1]/span[2]"
Title_xpath = "/html/body/div[1]/div[2]/div[2]/div/div[4]/div/table/thead/tr/th[2]"

#-------------------------------------------------------------------------

chrome_driver_path = ChromeDriverManager().install()
chrome_options = Options()
chrome_options.add_argument("--headless")  # 设置为无头模式
service = Service(chrome_driver_path)
driver = webdriver.Chrome(service=service,options=chrome_options)

# 创建一个Workbook对象
wb = Workbook()
ws = wb.active
row = 1

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
for StockNumber in StockList:
    url = replace_substring(TableURL, "stockID", StockNumber)
    hyperlink =  replace_substring(urlHis, "stockID", StockNumber)

    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        write_to_excel(wb, TableHeadRep, row)
        row += 1
        #行情分析
        soup = BeautifulSoup(response.text, 'html.parser')

        stock_detail = soup.select_one(".profile-detail").text.strip().split()

        stock_name = soup.select_one('.stock-name').text.strip()
        stock_name = re.match(r'(.+?)(?=\()', stock_name).group(0)
        print(f'股票名称: {stock_name}')
        ws.cell(row, 1, f'{stock_name}')
        cell = ws.cell(row,1)
        cell.font = Font(name='Arial', size=14, bold=True)
        cell.hyperlink = hyperlink

        stock_price_text = soup.select_one('.stock-current').text.strip().lstrip("¥")
        stock_price =  String2Number(stock_price_text)
        #print(f'股价：{stock_price_text})
        ws.cell(row, 2,stock_price)
        cell = ws.cell(row,2)
        cell.number_format = '"¥"#,##0.00'

        stock_change = soup.select_one('.stock-change').text.strip().split()[-1]
        #print(f'涨跌:{stock_change}')
        ws.cell(row, 3,f'{stock_change}')

        # 提取表格中的所有数据
        data_dict = {}
        current_key = None
        for td in soup.select('.quote-info td'):
            text = td.text.strip()
            if '：' in text:
                name, value = map(str.strip, text.split('：', 1))
                current_key = name
                data_dict[current_key] = float(value) if value.replace('.', '', 1).isdigit() else value

        colume = 4
        # 打印结果，每个键值对用换行符分隔
        for key, value in data_dict.items():
            if f'{key}' in TableHead:
              #print(f'{key}:{value}')
              ws.cell(row,TableHead.index(f'{key}')+2,f'{value}')
              colume +=1
              # 写入数据到单元格

        #价值分析
        url = replace_substring(urlDetail, "stockID", StockNumber)
        driver.get(url)
        checkbox = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, TongbeCheck_xpath)))
        if not checkbox.is_selected():
            checkbox.click()

        row += 1
        write_to_excel(wb, TableValue, row)
        ws.cell(row,1,StockNumber)
        cell =  ws.cell(row,1)
        cell.font = Font(name='Arial', size=14, bold=True)
        cell.hyperlink = replace_substring(TableURL, "stockID", StockNumber)

        row += 1
        ws.cell(row, 1, "当季报")
        colume =2
        for xpath in XpathTable:
            element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
            ws.cell(row, colume, element.text)
            colume += 1
            #print( element.text,'\n')

        button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, year_button_xpath)))
        button.click()
        row += 1
        ws.cell(row, 1, "去年报")
        colume = 2
        for xpath in XpathTable:
            element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
            ws.cell(row, colume, element.text)
            colume += 1
            #print( element.text,'\n')

        cell = ws.cell(row,2)
        if String2Number(cell.value)>0.2 : #ROE
            cell.font = Font(name='Arial', size=14, bold=True)
            cell.fill = PatternFill('solid', 'DDD9C4')

        cell =  ws.cell(row-3,5)
        if String2Number(cell.value)<15 : #PE
            cell.font = Font(name='Arial', size=14, bold=True)
            cell.fill = PatternFill('solid', 'DDD9C4')

        cell =  ws.cell(row-3,7)
        if String2Number(cell.value) > 0.04: #股息率
            cell.font = Font(name='Arial', size=14, bold=True)
            cell.fill = PatternFill('solid', 'DDD9C4')

        PEvsROE = String2Number(ws.cell(row,2).value)/String2Number(ws.cell(row-3,5).value) #ROE/PE
        ws.cell(row-3,13,PEvsROE)
        cell =  ws.cell(row-3,13)
        cell.number_format = '0.00%'
        cell.fill = PatternFill('solid','DDD9C4')
        cell.font = Font(name='Arial', size=14, bold=True)
        if PEvsROE > 0.01:
            cell.font = Font(name='Arial', size=14, bold=True,color='FF0000')

        #ROES = ws.cell(row-1,2).value

        GMY =String2Number( ws.cell(row,6).value)
        #GMS = ws.cell(row-1,6).value
        NMY = String2Number(ws.cell(row,8).value)
        #NMS = ws.cell(row-1,8).value
        cell =  ws.cell(row-3,14)

        cost = (GMY - NMY) /GMY
        ws.cell(row-3,14,cost)
        cell.number_format = '0.00%'
        cell.fill = PatternFill('solid','DDD9C4')
        cell.font = Font(name='Arial', size=14, bold=True)
        if cost<0.5:
            cell.font = Font(name='Arial', size=14, bold=True,color='FF0000')

        row+=1
        cell = ws.cell(row=row, column=1, value=stock_detail[0])
        start_cell = 'A{}'.format(row)
        end_cell = 'N{}'.format(row)
        ws.merge_cells(f'{start_cell}:{end_cell}')
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
        wb.save('Newestdata.xlsx')

        print('-----------------------------')
        row += 2  #下一个股票
    else:
        print(url,'请求失败，状态码:', response.status_code)

wb.save('Newestdata.xlsx')
convert_text_to_values("Newestdata.xlsx")
os.startfile("Newestdata.xlsx")